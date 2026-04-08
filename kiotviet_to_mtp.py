#!/usr/bin/env python3
# Venv setup: python3 -m venv venv ; source venv/bin/activate ; pip install xlrd xlwt openpyxl
from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable

import xlrd
import xlwt
from openpyxl import load_workbook

PRODUCT_HEADER_ALIASES = {
    "loai_hang": ["Loại hàng"],
    "nhom_hang": ["Nhóm hàng(3 Cấp)", "Nhóm hàng (3 Cấp)", "Nhóm hàng"],
    "ma_hang": ["Mã hàng"],
    "ten_hang": ["Tên hàng"],
    "gia_ban": ["Giá bán trước thuế", "Giá bán"],
    "gia_von": ["Giá vốn"],
    "ton_kho": ["Tồn kho"],
    "don_vi_tinh": ["ĐVT", "Đơn vị tính"],
}

CUSTOMER_HEADER_ALIASES = {
    "ma_khach_hang": ["Mã khách hàng"],
    "ten_khach_hang": ["Tên khách hàng"],
    "dien_thoai": ["Điện thoại"],
    "dia_chi": ["Địa chỉ"],
    "no_can_thu_hien_tai": ["Nợ cần thu hiện tại"],
}

PROVIDER_HEADER_ALIASES = {
    "ma_nha_cung_cap": ["Mã nhà cung cấp"],
    "ten_nha_cung_cap": ["Tên nhà cung cấp"],
    "email": ["Email"],
    "dien_thoai": ["Điện thoại"],
    "dia_chi": ["Địa chỉ"],
    "no_can_tra_hien_tai": ["Nợ cần trả hiện tại"],
}

SOURCE_PREFIXES = {
    "product": "DanhSachSanPham",
    "customer": "DanhSachKhachHang",
    "provider": "DanhSachNhaCungCap",
}

TEMPLATE_CANDIDATES = {
    "product_nganh": ["MTP_SP-NganhHang-LoaiHang.xls", "MTP-NganhHang-LoaiHang.xls"],
    "product_nhom": ["MTP_SP-NhomHang.xls", "MTP-NhomHang.xls"],
    "product_sanpham": ["MTP_SP-SanPham.xls", "MTP-SanPham.xls"],
    "product_tonkho": ["MTP_SP-TonKhoDauKy.xls", "Mau-TonKhoDauKy.xls"],
    "kh_ncc": ["MTP_KH-NCC.xls"],
    "kh_congno": ["MTP_KH-CongNoDauKy.xls"],
}


def clean_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        text = value.strip()
    else:
        text = str(value).strip()
    return re.sub(r"\s+", " ", text)


def normalize_header(value) -> str:
    text = clean_text(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", "", text.casefold())


def flatten_aliases(*alias_maps: dict[str, list[str]]) -> list[str]:
    return sorted(
        {
            alias
            for alias_map in alias_maps
            for aliases in alias_map.values()
            for alias in aliases
        }
    )


KNOWN_PRODUCT_HEADERS = flatten_aliases(PRODUCT_HEADER_ALIASES)
KNOWN_CUSTOMER_HEADERS = flatten_aliases(CUSTOMER_HEADER_ALIASES)
KNOWN_PROVIDER_HEADERS = flatten_aliases(PROVIDER_HEADER_ALIASES)


def resolve_columns(
    headers: list[str],
    aliases_map: dict[str, list[str]],
    source_label: str,
) -> dict[str, int]:
    normalized_headers = {
        normalize_header(header): idx + 1
        for idx, header in enumerate(headers)
        if clean_text(header)
    }
    resolved: dict[str, int] = {}
    missing: list[str] = []

    for field, aliases in aliases_map.items():
        match = None
        for alias in aliases:
            match = normalized_headers.get(normalize_header(alias))
            if match is not None:
                break
        if match is None:
            missing.append(f"{field} ({', '.join(aliases)})")
            continue
        resolved[field] = match

    if missing:
        raise ValueError(f"Thiếu cột bắt buộc trong file {source_label}: " + "; ".join(missing))

    return resolved


def slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    ascii_text = re.sub(r"[^A-Za-z0-9]+", "-", ascii_text).strip("-").upper()
    return ascii_text or "ITEM"


def make_unique_code(base: str, used: set[str], prefix: str = "") -> str:
    candidate = f"{prefix}{base}"
    if candidate not in used:
        used.add(candidate)
        return candidate
    index = 2
    while True:
        candidate = f"{prefix}{base}-{index}"
        if candidate not in used:
            used.add(candidate)
            return candidate
        index += 1


def read_xlsx_headers(path: Path) -> list[str]:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb.active
        return [clean_text(c.value) for c in ws[1]]
    finally:
        wb.close()


def read_mapped_xlsx_rows(
    path: Path,
    aliases_map: dict[str, list[str]],
    source_label: str,
    key_fields: tuple[str, ...],
) -> tuple[list[str], list[dict[str, object]]]:
    wb = load_workbook(path, data_only=True)
    try:
        ws = wb.active
        headers = [clean_text(c.value) for c in ws[1]]
        columns = resolve_columns(headers, aliases_map, source_label)
        rows: list[dict[str, object]] = []

        for row_idx in range(2, ws.max_row + 1):
            row = {
                field: ws.cell(row_idx, col_idx).value
                for field, col_idx in columns.items()
            }
            if all(clean_text(row[field]) == "" for field in key_fields):
                continue
            rows.append(row)

        return headers, rows
    finally:
        wb.close()


def read_kiotviet_rows(path: Path) -> tuple[list[str], list[dict[str, object]]]:
    return read_mapped_xlsx_rows(
        path,
        PRODUCT_HEADER_ALIASES,
        "KiotViet sản phẩm",
        ("ma_hang", "ten_hang"),
    )


def read_customer_rows(path: Path) -> tuple[list[str], list[dict[str, object]]]:
    return read_mapped_xlsx_rows(
        path,
        CUSTOMER_HEADER_ALIASES,
        "KiotViet khách hàng",
        ("ma_khach_hang", "ten_khach_hang"),
    )


def read_provider_rows(path: Path) -> tuple[list[str], list[dict[str, object]]]:
    return read_mapped_xlsx_rows(
        path,
        PROVIDER_HEADER_ALIASES,
        "KiotViet nhà cung cấp",
        ("ma_nha_cung_cap", "ten_nha_cung_cap"),
    )


def read_xls_rows(path: Path) -> tuple[list[str], list[list[object]]]:
    book = xlrd.open_workbook(path)
    sheet = book.sheet_by_index(0)
    rows = [sheet.row_values(r) for r in range(sheet.nrows)]
    headers = [clean_text(v) for v in rows[0]] if rows else []
    return headers, rows[1:] if len(rows) > 1 else []


def write_xls(
    path: Path,
    sheet_name: str,
    headers: list[str],
    rows: Iterable[Iterable[object]],
) -> None:
    wb = xlwt.Workbook()
    ws = wb.add_sheet(sheet_name)
    for c, value in enumerate(headers):
        ws.write(0, c, value)
    for r, row in enumerate(rows, start=1):
        for c, value in enumerate(row):
            ws.write(r, c, value)
    wb.save(str(path))


def to_number(value):
    if value is None or clean_text(value) == "":
        return None
    if isinstance(value, (int, float)):
        return value
    text = clean_text(value).replace(",", "")
    try:
        num = Decimal(text)
    except InvalidOperation:
        return value
    if num == num.to_integral_value():
        return int(num)
    return float(num)


def to_number_or_default(value, default=0):
    number = to_number(value)
    if number is None:
        return default
    return number


def detect_source_type(path: Path, headers: list[str]) -> str:
    normalized_name = normalize_header(path.stem)
    for source_type, prefix in SOURCE_PREFIXES.items():
        if normalized_name.startswith(normalize_header(prefix)):
            return source_type

    normalized_headers = {normalize_header(header) for header in headers if clean_text(header)}
    if {
        normalize_header("Mã hàng"),
        normalize_header("Tên hàng"),
    }.issubset(normalized_headers):
        return "product"
    if {
        normalize_header("Mã khách hàng"),
        normalize_header("Tên khách hàng"),
    }.issubset(normalized_headers):
        return "customer"
    if {
        normalize_header("Mã nhà cung cấp"),
        normalize_header("Tên nhà cung cấp"),
    }.issubset(normalized_headers):
        return "provider"

    raise ValueError(
        f"Không nhận diện được loại file: {path.name}. "
        "Tên file nên bắt đầu bằng DanhSachSanPham, DanhSachKhachHang hoặc DanhSachNhaCungCap."
    )


def resolve_template_path(templates_dir: Path, candidates: list[str]) -> Path | None:
    for name in candidates:
        path = templates_dir / name
        if path.exists():
            return path
    return None


def normalize_row(row: list[object], width: int) -> list[object]:
    return list(row[:width]) + [""] * max(0, width - len(row))


def build_nganh_hang(loai_hang_values: list[str], existing_rows: list[list[object]]) -> list[list[object]]:
    result = []
    existing_names = set()
    used_codes = set()
    for row in existing_rows:
        name = clean_text(row[1]) if len(row) > 1 else ""
        code = clean_text(row[0]) if len(row) > 0 else ""
        if not any(clean_text(cell) for cell in row[:3]):
            continue
        if name:
            existing_names.add(name.casefold())
        if code:
            used_codes.add(code)
        result.append(normalize_row(row, 3))

    for value in loai_hang_values:
        name = clean_text(value)
        if not name or name.casefold() in existing_names:
            continue
        suggested = "MD" if not used_codes else slugify(name)[:20]
        code = make_unique_code(suggested, used_codes)
        result.append([code, name, ""])
        existing_names.add(name.casefold())
    return result


def build_nhom_hang(nhom_hang_values: list[str], existing_rows: list[list[object]]) -> list[list[object]]:
    result = []
    existing_names = set()
    used_codes = set()
    for row in existing_rows:
        name = clean_text(row[1]) if len(row) > 1 else ""
        code = clean_text(row[0]) if len(row) > 0 else ""
        if not any(clean_text(cell) for cell in row[:4]):
            continue
        if name:
            existing_names.add(name.casefold())
        if code:
            used_codes.add(code)
        normalized = normalize_row(row, 4)
        if clean_text(normalized[2]) == "":
            normalized[2] = "MD"
        result.append(normalized)

    for value in nhom_hang_values:
        name = clean_text(value)
        if not name or name.casefold() in existing_names:
            continue
        code = make_unique_code(slugify(name)[:20], used_codes, prefix="NH-")
        result.append([code, name, "MD", ""])
        existing_names.add(name.casefold())
    return result


def build_san_pham(template_path: Path, output_path: Path, kiotviet_rows: list[dict[str, object]]) -> int:
    headers, existing_rows = read_xls_rows(template_path)
    result = []
    existing_codes = set()

    for row in existing_rows:
        normalized = normalize_row(row, 31)
        ma_san_pham = clean_text(normalized[1]) if len(normalized) > 1 else ""
        ten_san_pham = clean_text(normalized[2]) if len(normalized) > 2 else ""
        dedupe_key = ma_san_pham.casefold() or ten_san_pham.casefold()
        if dedupe_key:
            existing_codes.add(dedupe_key)
        if any(clean_text(cell) for cell in normalized):
            result.append(normalized)

    added = 0
    for item in kiotviet_rows:
        ma_hang = clean_text(item["ma_hang"])
        ten_hang = clean_text(item["ten_hang"])
        if not ma_hang and not ten_hang:
            continue
        dedupe_key = ma_hang.casefold() or ten_hang.casefold()
        if dedupe_key in existing_codes:
            continue
        row = [""] * 31
        row[0] = clean_text(item["nhom_hang"])
        row[1] = ma_hang
        row[2] = ten_hang
        row[4] = clean_text(item["don_vi_tinh"])
        row[5] = to_number(item["gia_von"])
        row[6] = to_number(item["gia_ban"])
        row[7] = 0
        row[8] = 999999
        result.append(row)
        existing_codes.add(dedupe_key)
        added += 1

    write_xls(
        output_path,
        "Sheet1",
        headers or [
            "Nhóm hàng hóa",
            "Mã sản phẩm",
            "Tên sản phẩm",
            "Tên viết tắt sản phẩm",
            "Đơn vị tính",
            "Giá nhập",
            "Giá bán",
            "Số lượng tồn tối thiểu",
            "Số lượng tồn tối đa",
            "Ghi chú",
        ],
        result,
    )
    return added


def build_ton_kho_dau_ky(template_path: Path, output_path: Path, kiotviet_rows: list[dict[str, object]]) -> int:
    headers, existing_rows = read_xls_rows(template_path)
    result = []
    existing_codes = set()

    for row in existing_rows:
        normalized = normalize_row(row, 3)
        ma_san_pham = clean_text(normalized[0])
        if ma_san_pham:
            existing_codes.add(ma_san_pham.casefold())
        if any(clean_text(cell) for cell in normalized):
            result.append(normalized)

    added = 0
    for item in kiotviet_rows:
        ma_hang = clean_text(item["ma_hang"])
        if not ma_hang:
            continue
        dedupe_key = ma_hang.casefold()
        if dedupe_key in existing_codes:
            continue
        result.append([
            ma_hang,
            to_number(item["ton_kho"]),
            to_number(item["gia_von"]),
        ])
        existing_codes.add(dedupe_key)
        added += 1

    write_xls(
        output_path,
        "Sheet1",
        headers or ["Mã sản phẩm", "Số lượng đầu kỳ", "Đơn giá đầu kỳ"],
        result,
    )
    return added


def build_kh_ncc(
    template_path: Path,
    output_path: Path,
    customer_rows: list[dict[str, object]],
    provider_rows: list[dict[str, object]],
) -> int:
    headers, existing_rows = read_xls_rows(template_path)
    result = []

    for row in existing_rows:
        normalized = normalize_row(row, 13)
        if any(clean_text(cell) for cell in normalized):
            result.append(normalized)

    added = 0

    for item in customer_rows:
        row = [""] * 13
        row[0] = clean_text(item["ma_khach_hang"])
        row[1] = clean_text(item["ten_khach_hang"])
        row[2] = clean_text(item["dia_chi"])
        row[4] = "KL"
        row[5] = "TRUE"
        row[7] = clean_text(item["dien_thoai"])
        result.append(row)
        added += 1

    for item in provider_rows:
        row = [""] * 13
        row[0] = clean_text(item["ma_nha_cung_cap"])
        row[1] = clean_text(item["ten_nha_cung_cap"])
        row[2] = clean_text(item["dia_chi"])
        row[4] = "NCC"
        row[6] = "TRUE"
        row[7] = clean_text(item["dien_thoai"])
        row[8] = clean_text(item["email"])
        result.append(row)
        added += 1

    write_xls(
        output_path,
        "Sheet1",
        headers or [
            "Mã KH-NCC",
            "Tên KH-NCC",
            "Địa chỉ",
            "Nhân viên phụ trách",
            "Nhóm KH-NCC",
            "Khách hàng(x)",
            "Nhà cung cấp(x)",
            "Điện thoại",
            "Email",
            "Ghi chú",
            "Điểm đầu kỳ",
            "Mã số thuế",
            "Ngày sinh nhật",
        ],
        result,
    )
    return added


def build_kh_cong_no_dau_ky(
    template_path: Path,
    output_path: Path,
    customer_rows: list[dict[str, object]],
    provider_rows: list[dict[str, object]],
) -> int:
    headers, existing_rows = read_xls_rows(template_path)
    result = []

    for row in existing_rows:
        normalized = normalize_row(row, 4)
        if any(clean_text(cell) for cell in normalized):
            result.append(normalized)

    added = 0

    for item in customer_rows:
        row = [""] * 4
        row[0] = clean_text(item["ma_khach_hang"])
        row[2] = to_number_or_default(item["no_can_thu_hien_tai"], default=0)
        row[3] = 0
        result.append(row)
        added += 1

    for item in provider_rows:
        row = [""] * 4
        row[0] = clean_text(item["ma_nha_cung_cap"])
        row[2] = 0
        row[3] = to_number_or_default(item["no_can_tra_hien_tai"], default=0)
        result.append(row)
        added += 1

    write_xls(
        output_path,
        "Sheet1",
        headers or [
            "Mã KH - NCC(*)",
            "Tên KH - NCC",
            "Số tiền phải thu (*)",
            "Số tiền phải trả (*)",
        ],
        result,
    )
    return added


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Chuyển dữ liệu sản phẩm, khách hàng và nhà cung cấp từ Excel KiotViet sang bộ file mẫu MTP."
    )
    parser.add_argument(
        "--kiotviet",
        required=True,
        nargs="+",
        type=Path,
        help="Một hoặc nhiều file Excel KiotViet .xlsx",
    )
    parser.add_argument(
        "--outdir",
        type=Path,
        default=Path(__file__).resolve().parent / "output",
        help="Thư mục xuất file (mặc định: ./output)",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    project_dir = Path(__file__).resolve().parent
    templates_dir = project_dir / "templates"

    product_rows: list[dict[str, object]] = []
    customer_rows: list[dict[str, object]] = []
    provider_rows: list[dict[str, object]] = []
    source_counts = {"product": 0, "customer": 0, "provider": 0}

    for source_path in args.kiotviet:
        if not source_path.exists():
            print(f"Không tìm thấy file: {source_path}", file=sys.stderr)
            return 1

        source_type = "unknown"
        try:
            headers = read_xlsx_headers(source_path)
            source_type = detect_source_type(source_path, headers)
            if source_type == "product":
                _, rows = read_kiotviet_rows(source_path)
                product_rows.extend(rows)
            elif source_type == "customer":
                _, rows = read_customer_rows(source_path)
                customer_rows.extend(rows)
            else:
                _, rows = read_provider_rows(source_path)
                provider_rows.extend(rows)
            source_counts[source_type] += 1
        except ValueError as exc:
            print(str(exc), file=sys.stderr)
            if source_type == "product":
                supported_headers = KNOWN_PRODUCT_HEADERS
            elif source_type == "customer":
                supported_headers = KNOWN_CUSTOMER_HEADERS
            elif source_type == "provider":
                supported_headers = KNOWN_PROVIDER_HEADERS
            else:
                supported_headers = flatten_aliases(
                    PRODUCT_HEADER_ALIASES,
                    CUSTOMER_HEADER_ALIASES,
                    PROVIDER_HEADER_ALIASES,
                )
            print(
                "Các tiêu đề đang hỗ trợ: " + ", ".join(supported_headers),
                file=sys.stderr,
            )
            return 1

    needs_product_outputs = bool(product_rows)
    needs_partner_outputs = bool(customer_rows or provider_rows)

    missing_templates: list[str] = []
    resolved_templates: dict[str, Path] = {}

    for key, candidates in TEMPLATE_CANDIDATES.items():
        if key.startswith("product_") and not needs_product_outputs:
            continue
        if key.startswith("kh_") and not needs_partner_outputs:
            continue
        path = resolve_template_path(templates_dir, candidates)
        if path is None:
            missing_templates.append("/".join(candidates))
            continue
        resolved_templates[key] = path

    if missing_templates:
        print(
            "Thiếu file template trong thư mục templates/: "
            + ", ".join(missing_templates),
            file=sys.stderr,
        )
        return 1

    args.outdir.mkdir(parents=True, exist_ok=True)

    if needs_product_outputs:
        mtp_nganh = resolved_templates["product_nganh"]
        mtp_nhom = resolved_templates["product_nhom"]
        mtp_sanpham = resolved_templates["product_sanpham"]
        ton_kho_dau_ky = resolved_templates["product_tonkho"]

        loai_hang_values = [clean_text(r["loai_hang"]) for r in product_rows]
        nhom_hang_values = [clean_text(r["nhom_hang"]) for r in product_rows]
        nganh_headers, nganh_existing = read_xls_rows(mtp_nganh)
        nhom_headers, nhom_existing = read_xls_rows(mtp_nhom)
        nganh_rows = build_nganh_hang(loai_hang_values, nganh_existing)
        nhom_rows = build_nhom_hang(nhom_hang_values, nhom_existing)

        out_nganh = args.outdir / mtp_nganh.name
        out_nhom = args.outdir / mtp_nhom.name
        out_sanpham = args.outdir / mtp_sanpham.name
        out_ton_kho = args.outdir / ton_kho_dau_ky.name

        write_xls(
            out_nganh,
            "Sheet2",
            nganh_headers or ["Mã ngành", "Tên ngành", "Ghi chú"],
            nganh_rows,
        )
        write_xls(
            out_nhom,
            "Sheet2",
            nhom_headers or ["Mã nhóm", "Tên nhóm", "Ngành hàng", "Ghi chú"],
            nhom_rows,
        )
        san_pham_added = build_san_pham(mtp_sanpham, out_sanpham, product_rows)
        ton_kho_added = build_ton_kho_dau_ky(ton_kho_dau_ky, out_ton_kho, product_rows)

        print(f"Đã xuất: {out_nganh}")
        print(f"Đã xuất: {out_nhom}")
        print(f"Đã xuất: {out_sanpham}")
        print(f"Đã xuất: {out_ton_kho}")
        print(f"Số dòng sản phẩm thêm mới: {san_pham_added}")
        print(f"Số dòng tồn kho đầu kỳ thêm mới: {ton_kho_added}")
        print(f"Số ngành hàng duy nhất: {len(nganh_rows)}")
        print(f"Số nhóm hàng duy nhất: {len(nhom_rows)}")
        print(
            "Ghi chú: cột C của MTP_SP-NhomHang được đặt cố định là 'MD'; "
            "cột H/I của MTP_SP-SanPham lần lượt là 0 và 999999."
        )

    if needs_partner_outputs:
        kh_ncc_template = resolved_templates["kh_ncc"]
        kh_congno_template = resolved_templates["kh_congno"]
        out_kh_ncc = args.outdir / kh_ncc_template.name
        out_kh_congno = args.outdir / kh_congno_template.name
        kh_ncc_added = build_kh_ncc(kh_ncc_template, out_kh_ncc, customer_rows, provider_rows)
        kh_congno_added = build_kh_cong_no_dau_ky(
            kh_congno_template,
            out_kh_congno,
            customer_rows,
            provider_rows,
        )

        print(f"Đã xuất: {out_kh_ncc}")
        print(f"Đã xuất: {out_kh_congno}")
        print(f"Số dòng KH/NCC thêm mới: {kh_ncc_added}")
        print(f"Số dòng công nợ đầu kỳ KH/NCC thêm mới: {kh_congno_added}")
        print(
            "Ghi chú: KH mặc định dùng nhóm 'KL' và cột Khách hàng(x) = TRUE; "
            "NCC dùng nhóm 'NCC' và cột Nhà cung cấp(x) = TRUE."
        )

    print(
        "Đã xử lý file nguồn: "
        f"{source_counts['product']} sản phẩm, "
        f"{source_counts['customer']} khách hàng, "
        f"{source_counts['provider']} nhà cung cấp."
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
