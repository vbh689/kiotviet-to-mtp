"""Excel readers and writers for KiotViet inputs and MTP templates."""

from __future__ import annotations

from pathlib import Path
from typing import Iterable, Mapping

import xlrd
import xlwt
from openpyxl import load_workbook

from .kv_config import (
    CUSTOMER_HEADER_ALIASES,
    PRODUCT_HEADER_ALIASES,
    PRODUCT_OPTIONAL_HEADER_ALIASES,
    PROVIDER_HEADER_ALIASES,
)
from .kv_utils import clean_text, normalize_header


def resolve_columns(
    headers: list[str],
    aliases_map: dict[str, list[str]],
    source_label: str,
    explicit_columns: Mapping[str, int] | None = None,
) -> dict[str, int]:
    if explicit_columns is not None:
        return validate_explicit_columns(headers, aliases_map, source_label, explicit_columns)

    # Build a "field name -> column number" mapping by matching normalized headers
    # against the alias lists above.
    resolved = resolve_alias_columns(headers, aliases_map)
    missing = [
        f"{field} ({', '.join(aliases)})"
        for field, aliases in aliases_map.items()
        if field not in resolved
    ]

    if missing:
        raise ValueError(f"Thiếu cột bắt buộc trong file {source_label}: " + "; ".join(missing))

    return resolved


def resolve_alias_columns(
    headers: list[str],
    aliases_map: dict[str, list[str]],
) -> dict[str, int]:
    # Return only fields that can be matched by aliases. GUI callers use this to
    # preselect defaults without treating missing headers as fatal yet.
    normalized_headers = {
        normalize_header(header): idx + 1
        for idx, header in enumerate(headers)
        if clean_text(header)
    }
    resolved: dict[str, int] = {}

    for field, aliases in aliases_map.items():
        match = None
        for alias in aliases:
            match = normalized_headers.get(normalize_header(alias))
            if match is not None:
                break
        if match is not None:
            resolved[field] = match

    return resolved


def validate_explicit_columns(
    headers: list[str],
    aliases_map: dict[str, list[str]],
    source_label: str,
    explicit_columns: Mapping[str, int],
) -> dict[str, int]:
    missing = [field for field in aliases_map if field not in explicit_columns]
    if missing:
        raise ValueError(
            f"Thiếu mapping bắt buộc trong file {source_label}: " + ", ".join(missing)
        )

    invalid: list[str] = []
    max_column = len(headers)
    resolved: dict[str, int] = {}

    for field in aliases_map:
        col_idx = explicit_columns[field]
        if isinstance(col_idx, bool) or not isinstance(col_idx, int):
            invalid.append(f"{field}={col_idx!r} không phải số cột")
            continue
        if col_idx < 1 or col_idx > max_column:
            invalid.append(f"{field}=cột {col_idx} ngoài phạm vi 1-{max_column}")
            continue
        resolved[field] = col_idx

    if invalid:
        raise ValueError(
            f"Mapping cột không hợp lệ trong file {source_label}: " + "; ".join(invalid)
        )

    return resolved


def read_excel_headers(path: Path) -> list[str]:
    # Read only the first row to detect the source type quickly.
    if path.suffix.lower() == '.xls':
        book = xlrd.open_workbook(path)
        sheet = book.sheet_by_index(0)
        return [clean_text(c) for c in sheet.row_values(0)] if sheet.nrows > 0 else []

    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb.active
        return [clean_text(c.value) for c in ws[1]]
    finally:
        wb.close()


def read_mapped_excel_rows(
    path: Path,
    aliases_map: dict[str, list[str]],
    source_label: str,
    key_fields: tuple[str, ...],
    column_mapping: Mapping[str, int] | None = None,
    optional_aliases_map: dict[str, list[str]] | None = None,
) -> tuple[list[str], list[dict[str, object]]]:
    # Read a KiotViet sheet and return rows keyed by our internal field names.
    # Blank data rows are skipped based on the important identifying columns.
    is_xls = path.suffix.lower() == '.xls'
    if is_xls:
        book = xlrd.open_workbook(path)
        sheet = book.sheet_by_index(0)
        headers = [clean_text(c) for c in sheet.row_values(0)] if sheet.nrows > 0 else []
        wb = None
        ws = None
    else:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        headers = [clean_text(c.value) for c in ws[1]]

    try:
        columns = resolve_columns(headers, aliases_map, source_label, column_mapping)
        optional_columns = {}
        if optional_aliases_map:
            if column_mapping:
                for field in optional_aliases_map:
                    if field in column_mapping:
                        col_idx = column_mapping[field]
                        if isinstance(col_idx, int) and 1 <= col_idx <= len(headers):
                            optional_columns[field] = col_idx
            else:
                optional_columns = resolve_alias_columns(headers, optional_aliases_map)

        rows: list[dict[str, object]] = []

        if is_xls:
            for row_idx in range(1, sheet.nrows):
                row_data = sheet.row_values(row_idx)
                row = {
                    field: row_data[col_idx - 1] if (col_idx - 1) < len(row_data) else None
                    for field, col_idx in columns.items()
                }
                if all(clean_text(row[field]) == "" for field in key_fields):
                    continue

                if optional_aliases_map:
                    for field in optional_aliases_map:
                        col_idx = optional_columns.get(field)
                        row[field] = row_data[col_idx - 1] if col_idx and (col_idx - 1) < len(row_data) else None

                rows.append(row)
        else:
            for row_idx in range(2, ws.max_row + 1):
                row = {
                    field: ws.cell(row_idx, col_idx).value
                    for field, col_idx in columns.items()
                }
                if all(clean_text(row[field]) == "" for field in key_fields):
                    continue

                if optional_aliases_map:
                    for field in optional_aliases_map:
                        col_idx = optional_columns.get(field)
                        row[field] = ws.cell(row_idx, col_idx).value if col_idx else None

                rows.append(row)

        return headers, rows
    finally:
        if wb:
            wb.close()


def read_kiotviet_rows(
    path: Path,
    column_mapping: Mapping[str, int] | None = None,
) -> tuple[list[str], list[dict[str, object]]]:
    # Thin wrappers keep the per-source configuration close to the call site.
    return read_mapped_excel_rows(
        path,
        PRODUCT_HEADER_ALIASES,
        "KiotViet sản phẩm",
        ("ma_hang", "ten_hang"),
        column_mapping,
        PRODUCT_OPTIONAL_HEADER_ALIASES,
    )


def read_customer_rows(
    path: Path,
    column_mapping: Mapping[str, int] | None = None,
) -> tuple[list[str], list[dict[str, object]]]:
    return read_mapped_excel_rows(
        path,
        CUSTOMER_HEADER_ALIASES,
        "KiotViet khách hàng",
        ("ma_khach_hang", "ten_khach_hang"),
        column_mapping,
    )


def read_provider_rows(
    path: Path,
    column_mapping: Mapping[str, int] | None = None,
) -> tuple[list[str], list[dict[str, object]]]:
    return read_mapped_excel_rows(
        path,
        PROVIDER_HEADER_ALIASES,
        "KiotViet nhà cung cấp",
        ("ma_nha_cung_cap", "ten_nha_cung_cap"),
        column_mapping,
    )


def read_xls_rows(path: Path) -> tuple[list[str], list[list[object]]]:
    # Template files are .xls, so they are handled with xlrd instead of openpyxl.
    book = xlrd.open_workbook(path)
    sheet = book.sheet_by_index(0)
    rows = [sheet.row_values(r) for r in range(sheet.nrows)]
    headers = [clean_text(v) for v in rows[0]] if rows else []
    return headers, rows[1:] if len(rows) > 1 else []


def write_xls(
    path: Path,
    sheet_name: str,
    headers: list[str],
    rows: Iterable[Iterable[object]],
) -> None:
    # Write a simple .xls file from headers plus row data.
    wb = xlwt.Workbook()
    ws = wb.add_sheet(sheet_name)
    for c, value in enumerate(headers):
        ws.write(0, c, value)
    for r, row in enumerate(rows, start=1):
        for c, value in enumerate(row):
            ws.write(r, c, value)
    wb.save(str(path))
